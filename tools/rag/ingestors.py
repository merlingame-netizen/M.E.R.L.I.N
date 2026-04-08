"""RAG Ingestors — parse various file types into indexable chunks.

Each ingestor returns a list of dicts: {text, id, metadata}.
"""

from __future__ import annotations

import csv
import logging
import os
import re
from pathlib import Path
from typing import Any

from rag.engine import make_chunk_id

logger = logging.getLogger("rag.ingestors")


# ── Markdown Ingestor ───────────────────────────────────────────────────────


def ingest_markdown(file_path: Path, source_type: str = "docs") -> list[dict[str, Any]]:
    """Split markdown by H2/H3 headings into chunks."""
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        logger.warning("Cannot read %s: %s", file_path, e)
        return []

    if not text.strip():
        return []

    # Extract frontmatter if present
    frontmatter = ""
    body = text
    if text.startswith("---"):
        end = text.find("---", 3)
        if end != -1:
            frontmatter = text[3:end].strip()
            body = text[end + 3 :].strip()

    # Split by headings (H2 or H3)
    sections = re.split(r"(?=^#{2,3}\s)", body, flags=re.MULTILINE)
    chunks: list[dict[str, Any]] = []
    src = str(file_path)

    for i, section in enumerate(sections):
        section = section.strip()
        if not section:
            continue

        # Extract heading title if present
        heading_match = re.match(r"^(#{2,3})\s+(.+)", section)
        heading = heading_match.group(2).strip() if heading_match else ""

        # Skip tiny sections (< 20 chars of content)
        content_text = re.sub(r"^#{2,3}\s+.+\n?", "", section).strip()
        if len(content_text) < 20 and not heading:
            continue

        chunk_text = section
        # Prepend frontmatter context to first chunk only
        if i == 0 and frontmatter:
            chunk_text = f"[{frontmatter}]\n\n{section}"

        chunks.append({
            "text": chunk_text[:4000],  # Cap at ~1000 tokens
            "id": make_chunk_id(src, i),
            "metadata": {
                "source_path": src,
                "source_type": source_type,
                "chunk_index": i,
                "heading": heading[:200],
                "file_name": file_path.name,
                "last_modified": file_path.stat().st_mtime,
            },
        })

    # If no headings found, index as single chunk
    if not chunks and len(body.strip()) >= 20:
        chunks.append({
            "text": body[:4000],
            "id": make_chunk_id(src, 0),
            "metadata": {
                "source_path": src,
                "source_type": source_type,
                "chunk_index": 0,
                "heading": file_path.stem,
                "file_name": file_path.name,
                "last_modified": file_path.stat().st_mtime,
            },
        })

    return chunks


# ── GDScript Ingestor ───────────────────────────────────────────────────────


def ingest_gdscript(file_path: Path) -> list[dict[str, Any]]:
    """Split GDScript by function definitions."""
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        logger.warning("Cannot read %s: %s", file_path, e)
        return []

    if not text.strip() or len(text) < 30:
        return []

    src = str(file_path)
    chunks: list[dict[str, Any]] = []

    # Split by func/static func definitions
    parts = re.split(r"(?=^(?:static\s+)?func\s+\w+)", text, flags=re.MULTILINE)

    for i, part in enumerate(parts):
        part = part.strip()
        if not part or len(part) < 20:
            continue

        # Extract function name
        func_match = re.match(r"(?:static\s+)?func\s+(\w+)", part)
        func_name = func_match.group(1) if func_match else ""

        chunks.append({
            "text": part[:4000],
            "id": make_chunk_id(src, i),
            "metadata": {
                "source_path": src,
                "source_type": "code",
                "chunk_index": i,
                "heading": func_name or f"header_{file_path.stem}",
                "file_name": file_path.name,
                "last_modified": file_path.stat().st_mtime,
            },
        })

    return chunks


# ── Excel/CSV Ingestor ──────────────────────────────────────────────────────


MAX_CHUNKS_PER_FILE = 50  # Cap to prevent bloated VOC Data indexing
MAX_SHEETS_PER_FILE = 3   # Only index first 3 sheets
MAX_ROWS_PER_SHEET = 500  # Sample first 500 data rows per sheet


def ingest_excel(file_path: Path) -> list[dict[str, Any]]:
    """Index Excel file — headers + sampled batches of rows (capped)."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed, skipping %s", file_path)
        return []

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Cannot read %s: %s", file_path, e)
        return []

    src = str(file_path)
    chunks: list[dict[str, Any]] = []
    chunk_idx = 0

    for sheet_name in wb.sheetnames[:MAX_SHEETS_PER_FILE]:
        ws = wb[sheet_name]
        rows = []
        for j, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if j >= MAX_ROWS_PER_SHEET:
                break
        if not rows:
            continue

        headers = [str(c) if c is not None else "" for c in rows[0]]
        header_line = " | ".join(h for h in headers if h)

        # Batch rows (100 per chunk, larger batches = fewer chunks)
        batch_size = 100
        data_rows = rows[1:]
        for batch_start in range(0, max(len(data_rows), 1), batch_size):
            if chunk_idx >= MAX_CHUNKS_PER_FILE:
                break
            batch = data_rows[batch_start : batch_start + batch_size]
            lines = [header_line]
            for row in batch:
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                if line.replace("|", "").strip():
                    lines.append(line)

            if len(lines) <= 1:
                continue

            text = f"[Sheet: {sheet_name}]\n" + "\n".join(lines)
            chunks.append({
                "text": text[:4000],
                "id": make_chunk_id(src, chunk_idx),
                "metadata": {
                    "source_path": src,
                    "source_type": "voc_data",
                    "chunk_index": chunk_idx,
                    "heading": f"{file_path.stem} / {sheet_name}",
                    "file_name": file_path.name,
                    "sheet": sheet_name,
                    "last_modified": file_path.stat().st_mtime,
                },
            })
            chunk_idx += 1

    try:
        wb.close()
    except Exception:
        pass
    return chunks


def ingest_csv(file_path: Path) -> list[dict[str, Any]]:
    """Index CSV file — headers + batches of rows."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as e:
        logger.warning("Cannot read %s: %s", file_path, e)
        return []

    if not rows:
        return []

    src = str(file_path)
    chunks: list[dict[str, Any]] = []
    headers = rows[0]
    header_line = " | ".join(headers)

    batch_size = 30
    data_rows = rows[1:]
    for i, batch_start in enumerate(range(0, max(len(data_rows), 1), batch_size)):
        batch = data_rows[batch_start : batch_start + batch_size]
        lines = [header_line]
        for row in batch:
            line = " | ".join(row)
            if line.replace("|", "").strip():
                lines.append(line)

        if len(lines) <= 1:
            continue

        text = "\n".join(lines)
        chunks.append({
            "text": text[:4000],
            "id": make_chunk_id(src, i),
            "metadata": {
                "source_path": src,
                "source_type": "voc_data",
                "chunk_index": i,
                "heading": file_path.stem,
                "file_name": file_path.name,
                "last_modified": file_path.stat().st_mtime,
            },
        })

    return chunks


# ── Outlook Ingestor ────────────────────────────────────────────────────────


def ingest_outlook(days: int = 180) -> list[dict[str, Any]]:
    """Fetch and chunk Outlook mails from the last N days."""
    try:
        import win32com.client
    except ImportError:
        logger.warning("win32com not available, skipping Outlook")
        return []

    try:
        app = win32com.client.Dispatch("Outlook.Application")
        ns = app.GetNamespace("MAPI")
    except Exception as e:
        logger.warning("Cannot connect to Outlook: %s", e)
        return []

    chunks: list[dict[str, Any]] = []
    import datetime

    cutoff = datetime.datetime.now() - datetime.timedelta(days=days)
    cutoff_str = cutoff.strftime("%m/%d/%Y")

    for folder_name, folder_id in [("inbox", 6), ("sent", 5)]:
        try:
            folder = ns.GetDefaultFolder(folder_id)
            items = folder.Items
            items.Sort("[ReceivedTime]", True)
            items = items.Restrict(f"[ReceivedTime] >= '{cutoff_str}'")
        except Exception as e:
            logger.warning("Cannot access %s: %s", folder_name, e)
            continue

        for i, item in enumerate(items):
            try:
                subject = getattr(item, "Subject", "") or ""
                sender = getattr(item, "SenderName", "") or ""
                to = getattr(item, "To", "") or ""
                received = getattr(item, "ReceivedTime", None)
                body = getattr(item, "Body", "") or ""

                date_str = str(received)[:19] if received else ""

                text = f"Subject: {subject}\nFrom: {sender}\nTo: {to}\nDate: {date_str}\n\n{body[:2000]}"

                mail_id = make_chunk_id(f"outlook:{folder_name}:{date_str}:{subject}", 0)

                chunks.append({
                    "text": text[:4000],
                    "id": mail_id,
                    "metadata": {
                        "source_path": f"outlook/{folder_name}",
                        "source_type": "outlook",
                        "chunk_index": i,
                        "heading": subject[:200],
                        "file_name": f"{folder_name}/{subject[:80]}",
                        "from": sender[:100],
                        "to": to[:100],
                        "date": date_str,
                        "last_modified": 0.0,
                    },
                })
            except Exception:
                continue

    return chunks


# ── TXT Ingestor ────────────────────────────────────────────────────────────


def ingest_txt(file_path: Path, source_type: str = "voc_data") -> list[dict[str, Any]]:
    """Index a plain text file, splitting by paragraphs or fixed size."""
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return []

    if len(text.strip()) < 20:
        return []

    src = str(file_path)
    # Split by double newlines (paragraphs)
    paragraphs = re.split(r"\n\s*\n", text)
    chunks: list[dict[str, Any]] = []

    # Merge small paragraphs into ~1000 char chunks
    current = ""
    chunk_idx = 0
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(current) + len(para) > 2000 and current:
            chunks.append({
                "text": current[:4000],
                "id": make_chunk_id(src, chunk_idx),
                "metadata": {
                    "source_path": src,
                    "source_type": source_type,
                    "chunk_index": chunk_idx,
                    "heading": file_path.stem,
                    "file_name": file_path.name,
                    "last_modified": file_path.stat().st_mtime,
                },
            })
            chunk_idx += 1
            current = para
        else:
            current = f"{current}\n\n{para}" if current else para

    if current.strip():
        chunks.append({
            "text": current[:4000],
            "id": make_chunk_id(src, chunk_idx),
            "metadata": {
                "source_path": src,
                "source_type": source_type,
                "chunk_index": chunk_idx,
                "heading": file_path.stem,
                "file_name": file_path.name,
                "last_modified": file_path.stat().st_mtime,
            },
        })

    return chunks


# ── Source discovery ────────────────────────────────────────────────────────

# Paths are resolved at call time, not import time
_SOURCE_CONFIGS = {
    "memory": {
        "path": lambda: Path.home() / ".claude" / "projects" / "c--Users-PGNK2128-Godot-MCP" / "memory",
        "glob": "*.md",
        "ingestor": lambda f: ingest_markdown(f, "memory"),
        "collection": "memory",
    },
    "docs": {
        "path": lambda: Path("c:/Users/PGNK2128/Godot-MCP/docs"),
        "glob": "**/*.md",
        "ingestor": lambda f: ingest_markdown(f, "docs"),
        "collection": "docs",
    },
    "agents": {
        "path": lambda: Path("c:/Users/PGNK2128/Godot-MCP/.claude/agents"),
        "glob": "*.md",
        "ingestor": lambda f: ingest_markdown(f, "agents"),
        "collection": "agents",
    },
    "code": {
        "path": lambda: Path("c:/Users/PGNK2128/Godot-MCP"),
        "glob": "**/*.gd",
        "ingestor": ingest_gdscript,
        "collection": "code",
        "exclude": ["addons/gut/", "archive/", ".godot/"],
    },
    "voc": {
        "path": lambda: Path("C:/Users/PGNK2128/OneDrive - orange.com/Partage VOC/Data"),
        "collection": "voc_data",
        # Multi-type: handled in discover_files()
    },
    "outlook": {
        "collection": "outlook",
        # Special: no file glob, uses COM
    },
}


def discover_files(source_name: str) -> list[tuple[Path, Any]]:
    """Return list of (file_path, ingestor_fn) for a source."""
    config = _SOURCE_CONFIGS.get(source_name)
    if not config:
        return []

    if source_name == "outlook":
        return []  # Handled specially via ingest_outlook()

    base_path = config["path"]()
    if not base_path.exists():
        logger.warning("Source path not found: %s", base_path)
        return []

    excludes = config.get("exclude", [])

    if source_name == "voc":
        # Multi-type discovery
        results: list[tuple[Path, Any]] = []
        for f in base_path.rglob("*.xlsx"):
            if not any(ex in str(f) for ex in [".git", "__pycache__", "node_modules", "~$"]):
                results.append((f, ingest_excel))
        for f in base_path.rglob("*.csv"):
            if not any(ex in str(f) for ex in [".git", "__pycache__", "node_modules"]):
                results.append((f, ingest_csv))
        for f in base_path.rglob("*.txt"):
            if not any(ex in str(f) for ex in [".git", "__pycache__", "node_modules"]):
                results.append((f, lambda fp: ingest_txt(fp, "voc_data")))
        for f in base_path.rglob("*.md"):
            if not any(ex in str(f) for ex in [".git", "__pycache__", "node_modules"]):
                results.append((f, lambda fp: ingest_markdown(fp, "voc_data")))
        return results

    glob_pattern = config.get("glob", "**/*")
    ingestor = config.get("ingestor")
    results = []
    for f in base_path.rglob(glob_pattern.replace("**/", "")):
        if f.is_file() and not any(ex in str(f) for ex in excludes + [".git", "__pycache__", ".godot"]):
            results.append((f, ingestor))

    return results
